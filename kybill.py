import string
import logging
import xml.etree.ElementTree as ET
from pathlib import Path
from argparse import ArgumentParser, BooleanOptionalAction

from openpyxl import Workbook
from openpyxl.styles import Font


def main():
    logger = logging.getLogger(name=__name__)

    args = ArgumentParser()
    args.add_argument("source",
                      help="Filepath of source Invoice")
    args.add_argument("destination", nargs="?",
                      help="Filepath of destination report (default same as source but .xlsx)")
    args.add_argument("--detailed", default=False, action=BooleanOptionalAction,
                      help="Enabled detailed expense")
    args = args.parse_args()

    _balance_start_val, _balance_start_sign = "Баланс поч.", "знак балансу"
    _balance_end_val, _balance_end_sign = "Баланс кін.", "знак балансу."
    fields = {
        "Ос. номер": "Invoice/Customer/BillingAccount",
        "Номер тел.": "Invoice/Customer/CustomerPhone",
        "Тарифний план": "Invoice/Contract/ContractDetail/ContractType",
        "Початкова дата": "Invoice/Header/BillingPeriod/BeginDate",
        "Кінцева дата": "Invoice/Header/BillingPeriod/EndDate",
        _balance_start_val: "Invoice/InvoiceAmount/AmountDetail/BalBeginMonth",
        _balance_start_sign: "Invoice/InvoiceAmount/AmountDetail/BalBeginMonthText",
        _balance_end_val: "Invoice/InvoiceAmount/AmountDetail/BalEndMonth",
        _balance_end_sign: "Invoice/InvoiceAmount/AmountDetail/BalBeginEndText",
        "Поповн.": "Invoice/InvoiceAmount/AmountDetail/Payments/PaymentsBank",
        "Рекоменд": "Invoice/InvoiceAmount/AmountDetail/RecommendedPayment",
        "Деталізація": "Invoice/Summary/SummaryRow/RowDetail//Text",
        "Витрати": "Invoice/Summary/SummaryRow/RowDetail//AmountExclTax",
        "ПДВ": 'Invoice/Summary/SummaryRow/RowDetail/TaxAmount[@Type="VAT"]',
        "Пенс. фонд": 'Invoice/Summary/SummaryRow/RowDetail/TaxAmount[@Type="PF"]',
        "Разом": "Invoice/Summary/SummaryRow/RowDetail/Amount"
    }
    keys = list(fields.keys())
    vals = list(fields.values())
    _header_size = 1

    src_xml = args.source
    dst_xls = args.destination if args.destination else f"{Path(src_xml).stem}.xlsx"

    wb = Workbook()
    ws = wb.active

    make_header(ws, keys)
    hide_column(ws, keys.index(_balance_start_sign))
    hide_column(ws, keys.index(_balance_end_sign))

    for row, cols in enumerate(parse_xml(src_xml, vals, args.detailed)):
        if len(keys) != len(cols):
            logger.warning("Skipping Row %s: %s", row, cols)
            continue

        transform_balance(
            balance=cols[keys.index(_balance_start_val)],
            balance_sign=cols[keys.index(_balance_start_sign)]
        )
        transform_balance(
            balance=cols[keys.index(_balance_end_val)],
            balance_sign=cols[keys.index(_balance_end_sign)]
        )

        write_cells(ws, row=row+_header_size, cols=cols)

    wb.save(dst_xls)


def transform_balance(balance, balance_sign):
    if balance_sign[0].text == "заборгованiсть":
        balance[0].text = f"-{balance[0].text}"


def make_header(work_sheet, head_list: list):
    for i, col in enumerate(head_list):
        cell = work_sheet.cell(row=1, column=i+1, value=col)
        # make title bold
        cell.font = Font(bold=True)
        # resize column
        work_sheet.column_dimensions[cell.column_letter].width = len(cell.value) + 2
    # freeze title row
    work_sheet.freeze_panes = "A2"


def hide_column(work_sheet, index):
    column_letter = string.ascii_uppercase[index]
    work_sheet.column_dimensions[column_letter].hidden = True


def parse_xml(xml_file, tag_list: list, detailed: bool):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    for account in root:
        row = []
        for tag in tag_list:
            if detailed:
                row.append(account.findall(tag))
                continue
            row.append([account.find(tag)])
        yield row


def write_cells(ws, row: int, cols: list):
    row = ws.max_row+1
    for cell, values in enumerate(cols):
        for r, val in enumerate(values):
            try:
                ws.cell(row=row+r, column=cell+1, value=to_float_safe(val.text))
            except AttributeError:
                continue


def to_float_safe(val):
    try:
        return float(val)
    except (TypeError, ValueError):
        return val


if __name__ == "__main__":
    main()
